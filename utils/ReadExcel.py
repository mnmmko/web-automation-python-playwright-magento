import openpyxl

class ReadExcel:

    @staticmethod
    def read_sheet(sheet_name):
        file_path = "./Data/user.xlsx"
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        rows = sheet.max_row
        columns = sheet.max_column
        data = []
        for i in range(2, rows + 1):
            row_data = []
            for j in range(1, columns + 1):
                cell_value = sheet.cell(row=i, column=j).value
                row_data.append("" if cell_value is None else str(cell_value))
            data.append(row_data)
        return data
